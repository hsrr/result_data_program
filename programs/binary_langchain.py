#: -*- coding: utf-8 -*-
"""
Filename:    solubility_langchain
Author:      chenxin
Date:        2023/8/1
Description: 
"""
import os
import re
import time
import openai
import logging
from dotenv import load_dotenv

import random

import pandas as pd
# from langchain.chat_models import AzureChatOpenAI
from langchain.chat_models.openai import ChatOpenAI
from langchain.prompts import PromptTemplate
from langchain.chains import LLMChain
from langchain.prompts.few_shot import FewShotPromptTemplate
import traceback
from openpyxl import Workbook

from utils1 import construct_icl_pool, load_json
from utils1 import ProteinSimilarityExampleSelector

# if not load_dotenv('.env'):
#     print("Warning: .env file not found")

# openai.api_type = "azure"
# openai.api_base = os.getenv("OPENAI_ENDPOINT")
# openai.api_version = "2023-03-15-preview"
# openai.api_key = os.getenv("OPENAI_API_KEY")
# DEPLOYMENT_NAME = os.getenv("OPENAI_DEPLOYMENT_NAME", "GPT-35")
# MODEL_NAME = os.getenv("OPENAI_MODEL_NAME", "gpt-35-turbo")


openai_api_key_pool = [
    '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',
    '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',
    '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',
]
label2id = {
    "membrane-bound": 0, 
    "soluble": 1, 
    "unknown": -1,
}
id2label = {idx: label for label, idx in label2id.items()}


residue2id = {"GLY": 0, "ALA": 1, "SER": 2, "PRO": 3, "VAL": 4, "THR": 5, "CYS": 6, "ILE": 7, "LEU": 8,
                  "ASN": 9, "ASP": 10, "GLN": 11, "LYS": 12, "GLU": 13, "MET": 14, "HIS": 15, "PHE": 16,
                  "ARG": 17, "TYR": 18, "TRP": 19}
id2residue = {idx: residue for residue, idx in residue2id.items()}

residue_symbol2id = {"G": 0, "A": 1, "S": 2, "P": 3, "V": 4, "T": 5, "C": 6, "I": 7, "L": 8, "N": 9,
                     "D": 10, "Q": 11, "K": 12, "E": 13, "M": 14, "H": 15, "F": 16, "R": 17, "Y": 18, "W": 19}
id2residue_symbol = {idx: residue_symbol for residue_symbol, idx in residue_symbol2id.items()}


   
class BinaryLocalizationClassifier(object):

    def __init__(self, train_data: pd.DataFrame, num_samples_for_each_class=2, num_classes=None,
                 verbose=False, sampling_method="random",
                 selector_cache_dir=None, id2label_func=lambda x: "soluble" if x == 1 else "membrane-bound"):
        super(BinaryLocalizationClassifier, self).__init__()

        # prompt design is adapted from https://arxiv.org/abs/2305.18365

        # todo: add example selector to do few-shot learning
        self.general_template = """You are a biology expert who has been deeply involved in the field of protein research for years. Given the amino acid sequence of 
a protein, your task is to predict the property of proteins using your experienced biological Property Prediction knowledge.
        """

        self.task_specific_template = """Given an amino acid sequence delimited by triple backticks,
        please determine whether the protein represented by this amino acid sequence is membrane-bound or soluble. 
        The candidate locations contain membrane-bound and soluble.
        Please answer with (only) one of the two candidate locations. 
        """

        # question
        self.question_template = """
        Amino acid sequence: ```{sequence}```
        Location: 
        """

        computed_num_classes = len(train_data["target"].unique())
        if num_classes is None:
            num_classes = computed_num_classes
        else:
            assert num_classes == computed_num_classes

        # ICL (in-context learning)
        if sampling_method == "similar":
            examples = [{"sequence": sample["sequence"], "target": id2label_func(sample["target"])}
                        for _, sample in train_data.iterrows()]
            example_selector = ProteinSimilarityExampleSelector(
                examples=examples, data_dir=selector_cache_dir, k=num_samples_for_each_class * num_classes,
            )
            #examole 为训练集所有例子，selector_cache_dir为"/root/data/solubility",k=4
            example_prompt = PromptTemplate(
                input_variables=["sequence", "target"],
                template="Amino acid sequence: {sequence}\nLocation: {target}")
            self.prompt_template = FewShotPromptTemplate(
                example_selector=example_selector,
                example_prompt=example_prompt,
                prefix=self.task_specific_template,
                suffix=self.question_template,
                input_variables=["sequence"]
            )
            print("there is similar sample")
            print(self.prompt_template) 
        elif sampling_method == "random": # random sampling
            icl_pool = construct_icl_pool(train_data, num_samples_for_each_class)
            self.icl_template = ""
            for sample in icl_pool:
                sequence = sample["sequence"]
                target = sample["target"]
                target_str = id2label_func(target)
                sample_str = f"""
                Amino acid sequence: {sequence}
                Location: {target_str}
                """
                self.icl_template += sample_str
            self.template = self.task_specific_template + self.icl_template + self.question_template
            self.prompt_template = PromptTemplate.from_template(self.template)
        else:
            self.template = self.task_specific_template + self.question_template
            self.prompt_template = PromptTemplate.from_template(self.template)
            print("there is zero sample")
            print(self.prompt_template)
#         self.llm = AzureChatOpenAI(client=openai.ChatCompletion,
#                                    deployment_name=DEPLOYMENT_NAME,
#                                    model_name=MODEL_NAME,
#                                    request_timeout=6,
#                                    temperature=0,
#                                    max_retries=2,
#                                    max_tokens=10)
#         self.llm = ChatOpenAI(model="gpt-4",openai_api_key="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",max_tokens=10)
        self.llm = ChatOpenAI(model="gpt-3.5-turbo",openai_api_key=random.choice(openai_api_key_pool),max_tokens=10,temperature=0)
        self.chain = LLMChain(llm=self.llm, prompt=self.prompt_template, verbose=verbose)
        self.pattern = self.config_pattern()

    def __call__(self, sequence):
        response = ""
        label = ""
#         response = self.chain.run(sequence=sequence)
#         label = self.postprocess(response)
        try:
#             print(self.prompt_template)
#             print(len(self.prompt_template))
#             logging.warning(type(sequence))
#             logging.warning('seq {}'.format(sequence))
            response = self.chain.run(sequence=sequence)
#             print("reeeeeeeeeeeeeeeeeeeee")

#             logging.warning(response)
            label = self.postprocess(response)
#             print("laaaaaaaaaaaaaaaaaaaaaa")
            return {
                "response": response,
                "label": label,
            }
        except Exception as e:
#             traceback.print_exc()
            logging.warning(f"An error {e} occurred.")
            exit(-1)
            return {
                "response": response,
                "label": label,
            }
            """每次调用LLMChain的run方法时:

(1) 调用prompt_template的format方法,传入当前输入,渲染得到完整的prompt字符串。

(2) prompt_template的format内部会调用example_selector的select_examples方法,选择示例。

(3) 根据选择的示例,用example_prompt渲染每个示例。

(4) 将prefix、rendered examples和suffix拼接起来,得到完整的prompt。"""
    def postprocess(self, text):
        text = text.lower()
        results = re.findall(self.pattern, text)
        label = results[0] if len(results) > 0 else "unknown"
        if label == "soluble": return 1
        if label == "membrane-bound": return 0
        return -1
    
    def config_pattern(self):
        return re.compile(
            "(soluble|membrane-bound)")



from argparse import ArgumentParser


def parse_args():
    parser = ArgumentParser()
    parser.add_argument("--num_samples_for_each_class", type=int, default=2,
                        help="Number of samples for each class in-context learning (default: 2)")
    parser.add_argument("--sampling_method", type=str, default="random",
                        help="Sampling method (default: random)")

    args = parser.parse_args()
    return args

def canonicalize(sequence):
    new_sequence = []
    for aa in sequence:
        
        idx = residue_symbol2id[aa]
        residue = id2residue[idx]
        new_sequence.append(residue)
    new_sequence = " ".join(new_sequence)
    return new_sequence
 
    


def main(args):

    
    print(f"Args: {args}")
    dataset_name = "subcellular_localization_2"
    data_dir = f"/root/data/{dataset_name}"

    dfs = {}
    datasets = {}
    
    for split in ["train", "valid", "test"]:
        filepath = os.path.join(data_dir, f"{dataset_name}_{split}.json")
        split_data = load_json(filepath)
        split_df = pd.DataFrame(split_data, columns=["sequence", "target"])
        datasets[split] = split_data
        dfs[split] = split_df
#         print("solit_df11111111111111111111111")
#         print(split_df.head())
    all_df = pd.concat([dfs["train"], dfs["valid"], dfs["test"]], axis=0)
    test_df = dfs["test"]

#     print(f"Label distribution: ")
#     print(f"{all_df['target'].value_counts()}")

    clf = BinaryLocalizationClassifier(train_data=dfs["train"], num_samples_for_each_class=args.num_samples_for_each_class,
                               num_classes=None, verbose=False,
                               sampling_method=args.sampling_method, selector_cache_dir=data_dir,
                               id2label_func=lambda x: "soluble" if x == 1 else "membrane-bound")

    num_correct = 0
    num_total = 0
    num_unknown = 0
    test_0_df = test_df[test_df["target"] == 0].iloc[::]
    test_1_df = test_df[test_df["target"] == 1].iloc[::]
    test_sample_df = pd.concat([test_0_df, test_1_df], axis=0)
    # test_sample_df = test_sample_df.sample(frac=1.)
    


    # 创建一个新的Excel工作簿
    workbook = Workbook()

    # 获取默认的活动工作表
    worksheet = workbook.active

    # 初始化行数
    row = 1

    for sample_idx, sample in test_df[1672:].iterrows():
        sequence, target = sample["sequence"], sample["target"]
        try:
                can_sequence = canonicalize(sequence)
        except KeyError as s:
                print(f"idx:{sample_idx},{sequence}，{s}")
                print("erro")
                continue
        print(f"idx: {sample_idx}, sequence: {sequence}")
        worksheet.cell(row=row, column=1, value=f"idx: {sample_idx}, sequence: {sequence}")
        row+=1
        print(f"idx: {sample_idx}, target  : {target}/{id2label[target]}")
        worksheet.cell(row=row, column=1, value=f"idx: {sample_idx}, target  : {target}/{id2label[target]}")
        row+=1
        result = clf(sequence)
        label = result["label"]
#         response = result["response"]
#         print(1111111111)
#         print(label)
        print(f"idx: {sample_idx}, label : {label}/{id2label[label]}")
        worksheet.cell(row=row, column=1, value=f"idx: {sample_idx}, label : {label}/{id2label[label]}")
        row+=1
        num_correct += label == target 
        num_total += 1 
        if label == -1: 
            num_unknown += 1 
            print(f"idx: {sample_idx}, unknown protein : {sequence}")
            worksheet.cell(row=row, column=1, value=f"idx: {sample_idx}, unknown protein : {sequence}")
            row+=1
            print(f"idx: {sample_idx}, unknown response: {result['response']}")
            worksheet.cell(row=row, column=1, value=f"idx: {sample_idx}, unknown response: {result['response']}")
            row+=1

        print("")
        worksheet.cell(row=row, column=1, value="")
        row+=1
        time.sleep(5)
    
    accuracy = num_correct / num_total    
    print(f"correct:{num_correct},total:{num_total},unknow:{num_unknown}")
    worksheet.cell(row=row, column=1, value=f"correct:{num_correct},total:{num_total},unknow:{num_unknown}") 
    row+=1
    print(f"Accuracy: {accuracy: .4f} ...")
    worksheet.cell(row=row, column=1, value=f"Accuracy: {accuracy: .4f} ...") 
    workbook.save("binary_similar_3.5.xlsx")
    
#      for sample_idx, sample in test_sample_df.iterrows():
#         sequence, target = sample["sequence"], sample["target"]
#         print(f"sequence: {sequence}")
#         print(f"target  : {target}")
#         result = clf(sequence)
#         label = result["label"]
#         response = result["response"]
#         print(f"label   : {label}")
#         num_correct += label == target
#         num_total += 1
#         if label == -1:
#             num_unknown += 1
#             print(f"Unknown protein : {sequence}")
#             print(f"Unknown response: {response}")
#         print("")
#         time.sleep(5)

#     accuracy = num_correct / num_total
#     print(f"Accuracy: {accuracy: .4f} ...")

if __name__ == "__main__":
#     format_ = "%(levelname)s: %(asctime)s: %(filename)s:%(lineno)d * %(thread)d %(message)s"
#     datefmt="%Y-%m-%d-%H:%M:%S"
#     logging.basicConfig(level=logging.INFO, format=format_, datefmt=datefmt)
#     logging.root.setLevel(logging.DEBUG)
    args = parse_args()
    main(args)
